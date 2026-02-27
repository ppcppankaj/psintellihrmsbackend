"""Expense Views - API Endpoints"""

from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
from django_filters.rest_framework import DjangoFilterBackend

from apps.core.tenant_guards import OrganizationViewSetMixin
from apps.core.permissions import FilterByPermissionMixin, HasPermission
from .permissions import ExpensesTenantPermission
from apps.core.mixins import BulkImportExportMixin
from .filters import ExpenseCategoryFilter, ExpenseClaimFilter, ExpenseItemFilter, EmployeeAdvanceFilter
from .models import (
    ExpenseCategory, ExpenseClaim, ExpenseItem, EmployeeAdvance
)
from .serializers import (
    ExpenseCategorySerializer,
    ExpenseClaimListSerializer, ExpenseClaimDetailSerializer,
    CreateExpenseClaimSerializer, ApproveExpenseSerializer, ProcessPaymentSerializer,
    ExpenseItemSerializer,
    EmployeeAdvanceListSerializer, EmployeeAdvanceDetailSerializer,
    CreateAdvanceSerializer, ApproveAdvanceSerializer, DisburseAdvanceSerializer,
    BulkApproveRejectSerializer, ExportReportSerializer
)
from .services import ExpenseService, AdvanceService


class ExpenseCategoryViewSet(BulkImportExportMixin, OrganizationViewSetMixin, viewsets.ModelViewSet):
    """ViewSet for expense categories"""
    queryset = ExpenseCategory.objects.none()
    serializer_class = ExpenseCategorySerializer
    permission_classes = [IsAuthenticated, ExpensesTenantPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = ExpenseCategoryFilter
    search_fields = ['name', 'code']
    ordering_fields = ['name', 'created_at']
    ordering = ['name']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset


class ExpenseClaimViewSet(OrganizationViewSetMixin, FilterByPermissionMixin, viewsets.ModelViewSet):
    """
    ViewSet for expense claims.
    """
    queryset = ExpenseClaim.objects.none()
    permission_classes = [IsAuthenticated, ExpensesTenantPermission, HasPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = ExpenseClaimFilter
    search_fields = ['title', 'claim_number']
    ordering_fields = ['claim_date', 'total_claimed_amount', 'status', 'created_at']
    ordering = ['-claim_date']
    required_permissions = {
        'list': ['expenses.view'],
        'retrieve': ['expenses.view'],
        'create': ['expenses.create'],
    }
    scope_field = 'employee'
    permission_category = 'expenses'
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return ExpenseClaimDetailSerializer
        if self.action == 'create':
            return CreateExpenseClaimSerializer
        return ExpenseClaimListSerializer
    
    def get_employee(self, request):
        """Helper to get employee profile for current user"""
        if not hasattr(request, '_employee'):
            from apps.employees.models import Employee
            request._employee = Employee.objects.filter(user=request.user).first()
        return request._employee
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by status
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Filter by employee
        employee_id = self.request.query_params.get('employee')
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        
        # Filter by date range
        from_date = self.request.query_params.get('from_date')
        to_date = self.request.query_params.get('to_date')
        if from_date:
            queryset = queryset.filter(claim_date__gte=from_date)
        if to_date:
            queryset = queryset.filter(claim_date__lte=to_date)
        
        return queryset.select_related('employee', 'current_approver').prefetch_related('items')
    
    def create(self, request, *args, **kwargs):
        """Create expense claim with items"""
        serializer = CreateExpenseClaimSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        employee = self.get_employee(request)
        if not employee:
            return Response({'error': 'No employee profile'}, status=status.HTTP_400_BAD_REQUEST)
        
        claim = ExpenseService.create_expense_claim(
            employee=employee,
            title=serializer.validated_data['title'],
            description=serializer.validated_data.get('description', ''),
            claim_date=serializer.validated_data['claim_date'],
            expense_from=serializer.validated_data['expense_from'],
            expense_to=serializer.validated_data['expense_to'],
            items_data=serializer.validated_data['items']
        )
        
        return Response(
            ExpenseClaimDetailSerializer(claim).data,
            status=status.HTTP_201_CREATED
        )
    
    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Submit claim for approval"""
        claim = self.get_object()
        employee = self.get_employee(request)
        
        try:
            claim = ExpenseService.submit_claim(claim, employee)
            return Response(ExpenseClaimDetailSerializer(claim).data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve or reject expense claim"""
        claim = self.get_object()
        serializer = ApproveExpenseSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        approver = self.get_employee(request)
        
        try:
            claim = ExpenseService.process_approval(
                claim=claim,
                approver=approver,
                action=serializer.validated_data['action'],
                comments=serializer.validated_data.get('comments', ''),
                item_adjustments=serializer.validated_data.get('item_adjustments')
            )
            return Response(ExpenseClaimDetailSerializer(claim).data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def pay(self, request, pk=None):
        """Process payment for approved claim"""
        claim = self.get_object()
        serializer = ProcessPaymentSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        paid_by = self.get_employee(request)
        
        try:
            claim = ExpenseService.process_payment(
                claim=claim,
                paid_by=paid_by,
                amount=serializer.validated_data['amount'],
                payment_mode=serializer.validated_data['payment_mode'],
                payment_reference=serializer.validated_data.get('payment_reference', ''),
                advance_to_adjust=serializer.validated_data.get('adjust_advance_id')
            )
            return Response(ExpenseClaimDetailSerializer(claim).data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def my_claims(self, request):
        """Get current user's expense claims"""
        employee = self.get_employee(request)
        if not employee:
            return Response({'error': 'No employee profile'}, status=status.HTTP_404_NOT_FOUND)
        
        claims = ExpenseClaim.objects.filter(employee=employee).order_by('-claim_date')
        serializer = ExpenseClaimListSerializer(claims, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def pending_approvals(self, request):
        """Get claims pending current user's approval"""
        employee = self.get_employee(request)
        if not employee:
            return Response({'error': 'No employee profile'}, status=status.HTTP_404_NOT_FOUND)
        
        claims = ExpenseClaim.objects.filter(
            current_approver=employee,
            status__in=[ExpenseClaim.STATUS_SUBMITTED, ExpenseClaim.STATUS_PENDING_APPROVAL]
        ).order_by('-submitted_at' if hasattr(ExpenseClaim, 'submitted_at') else '-created_at')
        
        serializer = ExpenseClaimListSerializer(claims, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get expense summary for current user"""
        employee = self.get_employee(request)
        if not employee:
            return Response({'error': 'No employee profile'}, status=status.HTTP_404_NOT_FOUND)
        
        year = request.query_params.get('year')
        summary = ExpenseService.get_employee_expense_summary(employee, int(year) if year else None)
        return Response(summary)
    
    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Cancel a draft claim"""
        claim = self.get_object()
        
        if claim.status != ExpenseClaim.STATUS_DRAFT:
            return Response(
                {'error': 'Only draft claims can be cancelled'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        claim.status = ExpenseClaim.STATUS_CANCELLED
        claim.save(update_fields=['status'])
        
        return Response({'status': 'cancelled'})
    
    @action(detail=False, methods=['post'])
    def bulk_approve(self, request):
        """Bulk approve multiple expense claims"""
        serializer = BulkApproveRejectSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        approver = self.get_employee(request)
        if not approver:
            return Response({'error': 'No employee profile'}, status=status.HTTP_400_BAD_REQUEST)
        
        claim_ids = serializer.validated_data['claim_ids']
        comments = serializer.validated_data.get('comments', '')
        
        results = {'approved': [], 'failed': []}
        
        claims = ExpenseClaim.objects.filter(
            id__in=claim_ids,
            current_approver=approver,
            status__in=[ExpenseClaim.STATUS_SUBMITTED, ExpenseClaim.STATUS_PENDING_APPROVAL]
        )
        
        for claim in claims:
            try:
                ExpenseService.process_approval(
                    claim=claim,
                    approver=approver,
                    action='approve',
                    comments=comments
                )
                results['approved'].append(str(claim.id))
            except ValueError as e:
                results['failed'].append({'id': str(claim.id), 'error': str(e)})
        
        not_found = set(str(cid) for cid in claim_ids) - set(results['approved']) - set(f['id'] for f in results['failed'])
        for cid in not_found:
            results['failed'].append({'id': cid, 'error': 'Claim not found or not pending your approval'})
        
        return Response(results)
    
    @action(detail=False, methods=['post'])
    def bulk_reject(self, request):
        """Bulk reject multiple expense claims"""
        serializer = BulkApproveRejectSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        approver = self.get_employee(request)
        if not approver:
            return Response({'error': 'No employee profile'}, status=status.HTTP_400_BAD_REQUEST)
        
        claim_ids = serializer.validated_data['claim_ids']
        comments = serializer.validated_data.get('comments', '')
        
        if not comments:
            return Response(
                {'error': 'Comments required for rejection'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        results = {'rejected': [], 'failed': []}
        
        claims = ExpenseClaim.objects.filter(
            id__in=claim_ids,
            current_approver=approver,
            status__in=[ExpenseClaim.STATUS_SUBMITTED, ExpenseClaim.STATUS_PENDING_APPROVAL]
        )
        
        for claim in claims:
            try:
                ExpenseService.process_approval(
                    claim=claim,
                    approver=approver,
                    action='reject',
                    comments=comments
                )
                results['rejected'].append(str(claim.id))
            except ValueError as e:
                results['failed'].append({'id': str(claim.id), 'error': str(e)})
        
        not_found = set(str(cid) for cid in claim_ids) - set(results['rejected']) - set(f['id'] for f in results['failed'])
        for cid in not_found:
            results['failed'].append({'id': cid, 'error': 'Claim not found or not pending your approval'})
        
        return Response(results)
    
    @action(detail=False, methods=['get'])
    def export_report(self, request):
        """Export expense claims report as CSV or Excel"""
        import csv
        import io
        from django.http import HttpResponse
        
        serializer = ExportReportSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        
        export_format = serializer.validated_data.get('format', 'csv')
        from_date = serializer.validated_data.get('from_date')
        to_date = serializer.validated_data.get('to_date')
        status_filter = serializer.validated_data.get('status', 'all')
        employee_ids = serializer.validated_data.get('employee_ids', [])
        
        queryset = self.get_queryset()
        
        if from_date:
            queryset = queryset.filter(claim_date__gte=from_date)
        if to_date:
            queryset = queryset.filter(claim_date__lte=to_date)
        if status_filter and status_filter != 'all':
            queryset = queryset.filter(status=status_filter)
        if employee_ids:
            queryset = queryset.filter(employee_id__in=employee_ids)
        
        queryset = queryset.select_related('employee', 'approved_by').prefetch_related('items')
        
        headers = [
            'Claim Number', 'Employee ID', 'Employee Name', 'Title', 'Claim Date',
            'Expense From', 'Expense To', 'Total Claimed', 'Total Approved',
            'Total Paid', 'Status', 'Payment Status', 'Approved By', 'Approved At'
        ]
        
        rows = []
        for claim in queryset:
            rows.append([
                claim.claim_number,
                claim.employee.employee_id if claim.employee else '',
                claim.employee.full_name if claim.employee else '',
                claim.title,
                claim.claim_date.isoformat() if claim.claim_date else '',
                claim.expense_from.isoformat() if claim.expense_from else '',
                claim.expense_to.isoformat() if claim.expense_to else '',
                str(claim.total_claimed_amount),
                str(claim.total_approved_amount),
                str(claim.total_paid_amount),
                claim.get_status_display(),
                claim.get_payment_status_display(),
                claim.approved_by.full_name if claim.approved_by else '',
                claim.approved_at.isoformat() if claim.approved_at else ''
            ])
        
        if export_format == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            writer.writerows(rows)
            
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="expense_report.csv"'
            return response
        
        else:
            try:
                import openpyxl
                from openpyxl.utils import get_column_letter
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Expense Report"
                
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_num, value=header)
                
                for row_num, row_data in enumerate(rows, 2):
                    for col_num, cell_value in enumerate(row_data, 1):
                        ws.cell(row=row_num, column=col_num, value=cell_value)
                
                for col_num in range(1, len(headers) + 1):
                    ws.column_dimensions[get_column_letter(col_num)].width = 15
                
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = HttpResponse(
                    output.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename="expense_report.xlsx"'
                return response
                
            except ImportError:
                return Response(
                    {'error': 'Excel export requires openpyxl library'},
                    status=status.HTTP_501_NOT_IMPLEMENTED
                )


class ExpenseItemViewSet(OrganizationViewSetMixin, viewsets.ModelViewSet):
    """ViewSet for expense items"""
    queryset = ExpenseItem.objects.none()
    serializer_class = ExpenseItemSerializer
    permission_classes = [IsAuthenticated, ExpensesTenantPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = ExpenseItemFilter
    search_fields = ['description']
    ordering_fields = ['amount', 'created_at']
    ordering = ['created_at']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        claim_id = self.request.query_params.get('claim')
        if claim_id:
            queryset = queryset.filter(claim_id=claim_id)
        
        return queryset.select_related('claim', 'category')


class EmployeeAdvanceViewSet(OrganizationViewSetMixin, FilterByPermissionMixin, viewsets.ModelViewSet):
    """
    ViewSet for employee advances.
    """
    queryset = EmployeeAdvance.objects.none()
    permission_classes = [IsAuthenticated, ExpensesTenantPermission, HasPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = EmployeeAdvanceFilter
    search_fields = ['purpose']
    ordering_fields = ['advance_date', 'amount', 'status', 'created_at']
    ordering = ['-advance_date']
    required_permissions = {
        'list': ['expenses.view'],
        'retrieve': ['expenses.view'],
        'create': ['expenses.create'],
    }
    scope_field = 'employee'
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return EmployeeAdvanceDetailSerializer
        if self.action == 'create':
            return CreateAdvanceSerializer
        return EmployeeAdvanceListSerializer
    
    def get_employee(self, request):
        """Helper to get employee profile for current user"""
        if not hasattr(request, '_employee'):
            from apps.employees.models import Employee
            request._employee = Employee.objects.filter(user=request.user).first()
        return request._employee
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        employee_id = self.request.query_params.get('employee')
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        
        return queryset.select_related('employee', 'approved_by', 'disbursed_by')
    
    def create(self, request, *args, **kwargs):
        """Create advance request"""
        serializer = CreateAdvanceSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        employee = self.get_employee(request)
        if not employee:
            return Response({'error': 'No employee profile'}, status=status.HTTP_400_BAD_REQUEST)
        
        advance = AdvanceService.create_advance_request(
            employee=employee,
            purpose=serializer.validated_data['purpose'],
            advance_date=serializer.validated_data['advance_date'],
            amount=serializer.validated_data['amount'],
            settlement_type=serializer.validated_data['settlement_type'],
            deduction_start_month=serializer.validated_data.get('deduction_start_month'),
            monthly_deduction_amount=serializer.validated_data.get('monthly_deduction_amount')
        )
        
        return Response(
            EmployeeAdvanceDetailSerializer(advance).data,
            status=status.HTTP_201_CREATED
        )
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve or reject advance"""
        advance = self.get_object()
        serializer = ApproveAdvanceSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        approver = self.get_employee(request)
        
        try:
            advance = AdvanceService.approve_advance(
                advance=advance,
                approved_by=approver,
                action=serializer.validated_data['action'],
                comments=serializer.validated_data.get('comments', '')
            )
            return Response(EmployeeAdvanceDetailSerializer(advance).data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def disburse(self, request, pk=None):
        """Disburse approved advance"""
        advance = self.get_object()
        serializer = DisburseAdvanceSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        disbursed_by = self.get_employee(request)
        
        try:
            advance = AdvanceService.disburse_advance(
                advance=advance,
                disbursed_by=disbursed_by,
                disbursement_mode=serializer.validated_data['disbursement_mode'],
                disbursement_reference=serializer.validated_data.get('disbursement_reference', '')
            )
            return Response(EmployeeAdvanceDetailSerializer(advance).data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def my_advances(self, request):
        """Get current user's advances"""
        employee = self.get_employee(request)
        if not employee:
            return Response({'error': 'No employee profile'}, status=status.HTTP_404_NOT_FOUND)
        
        advances = EmployeeAdvance.objects.filter(employee=employee).order_by('-advance_date')
        serializer = EmployeeAdvanceListSerializer(advances, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def pending(self, request):
        """Get pending/unsettled advances for current user"""
        employee = self.get_employee(request)
        if not employee:
            return Response({'error': 'No employee profile'}, status=status.HTTP_404_NOT_FOUND)
        
        advances = AdvanceService.get_pending_advances(employee)
        serializer = EmployeeAdvanceListSerializer(advances, many=True)
        return Response(serializer.data)
